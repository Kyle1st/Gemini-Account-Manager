from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADERS = ["账号邮箱", "密码", "辅助邮箱", "TOTP密钥", "备注", "创建时间", "更新时间"]
FIELD_KEYS = ["email", "password", "recovery_email", "totp_secret", "notes", "created_at", "updated_at"]


def export_to_excel(accounts: list[dict], filepath: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, acc in enumerate(accounts, start=2):
        for col_idx, key in enumerate(FIELD_KEYS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=acc.get(key, ""))
            cell.border = thin_border

    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row in range(2, len(accounts) + 2):
            val = str(ws.cell(row=row, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    wb.save(filepath)


def import_from_excel(filepath: str) -> list[dict]:
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    accounts = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue
        acc = {}
        for i, key in enumerate(FIELD_KEYS):
            acc[key] = str(row[i]) if i < len(row) and row[i] is not None else ""
        accounts.append(acc)

    wb.close()
    return accounts
